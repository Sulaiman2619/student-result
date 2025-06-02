# students/views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from django.db.models import Q
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect,Http404,HttpResponseForbidden
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime  # Import the datetime module
import json
from urllib.parse import urlencode
from django.core.paginator import Paginator
from django.core.exceptions import PermissionDenied
from django.views.decorators.csrf import csrf_protect
from axes.helpers import get_client_ip_address
from django.utils.timezone import now
from axes.handlers.proxy import AxesProxyHandler
from axes.models import AccessAttempt
from django.conf import settings
from datetime import timedelta
import os
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib.colors import black
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter  # เพิ่มตรงนี้
from openpyxl.drawing.image import Image as XLImage
from urllib.parse import quote

pdfmetrics.registerFont(TTFont('THSarabunNew', 'static/fonts/THSarabunNew.ttf'))


def rotated_paragraph(text, width=40, height=100):
    """สร้างข้อความแนวตั้งด้วย Drawing"""
    d = Drawing(width, height)
    d.add(String(0, 0, text, fontName='THSarabunNew', fontSize=12, fillColor=black, textAnchor='start', angle=90))
    return d

def convert_to_thai_year(academic_year):
    try:
        # กรณีเป็นรูปแบบ 2023-2024
        if '-' in academic_year:
            parts = academic_year.split('-')
            return f"{int(parts[0]) + 543}-{int(parts[1]) + 543}"
        else:
            # กรณีเป็นปีเดียว เช่น 2023
            return str(int(academic_year) + 543)
    except:
        return academic_year  # ถ้าแปลงไม่ได้ก็คืนค่าเดิม

def delete_student(request, student_id):
    # Fetch the student by ID

    # Check if 'user_type' is in the session
    user_type = request.session.get('user_type')
    
    # If no user_type is found, redirect to login
    if not user_type:
        return redirect('login_view')  # Redirect to login if user_type is not in session

    # If user is not a teacher, redirect to home
    if user_type == 'student':
        return redirect('home')  # Redirect student back to home
    
    student = get_object_or_404(Student, id=student_id)
    # Soft delete the student by setting delete_status to 'deleted'
    student.delete_status = 'deleted'
    student.save()

    # Redirect back to the student report page
    return redirect('sp_student_report')  # Replace 'student_report' with the correct URL name for your student list page

def get_schools(request):
    schools = list(School.objects.values('id', 'name'))
    return JsonResponse(schools, safe=False)

def safe_int(value):
    """Convert a value to int or return 0 if it's not valid."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0

def add_school(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        school_name = data.get('name')
        if school_name:
            school, created = School.objects.get_or_create(name=school_name)
            return JsonResponse({'id': school.id, 'name': school.name})
        return JsonResponse({'error': 'Invalid data'}, status=400)
                            
def get_provinces(request):
    provinces = list(Province.objects.all().values('id', 'name'))
    return JsonResponse(provinces, safe=False)



# Get districts (amphoes) based on the selected province
def get_districts(request):
    province_id = request.GET.get('province_id')
    if not province_id:
        return JsonResponse({'error': 'Invalid province ID'}, status=400)
    
    districts = list(Amphoe.objects.filter(province_id=province_id).values('id', 'name'))
    return JsonResponse(districts, safe=False)

# Get subdistricts (tambons) based on the selected district
def get_subdistricts(request):
    district_id = request.GET.get('district_id')
    if not district_id:
        return JsonResponse({'error': 'Invalid district ID'}, status=400)
    
    subdistricts = list(Tambon.objects.filter(amphoe_id=district_id).values('id', 'name'))
    return JsonResponse(subdistricts, safe=False)

# Fetch zipcode based on subdistrict
def get_zipcode(request):
    subdistrict_id = request.GET.get('subdistrict_id')
    subdistrict = Tambon.objects.filter(id=subdistrict_id).first()
    if subdistrict:
        return JsonResponse({'zipcode': subdistrict.zipcode})
    return JsonResponse({'zipcode': ''})

@csrf_protect
def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    error_message = None

    # ✅ ล้าง record ที่หมดอายุแล้ว ก่อนเช็คบล็อก
    cutoff = now() - timedelta(minutes=getattr(settings, 'AXES_COOLOFF_TIME', 5))
    AccessAttempt.objects.filter(attempt_time__lt=cutoff).delete()

    # ✅ ตรวจสอบการบล็อกจาก django-axes และ ratelimit
    handler = AxesProxyHandler()
    blocked_axes = handler.is_locked(request)
    blocked_ratelimit = getattr(request, 'limited', False)
    blocked = blocked_axes or blocked_ratelimit

    # แจ้งเตือนถ้าถูกบล็อก
    if blocked:
        error_message = 'คุณถูกบล็อกจากระบบชั่วคราว กรุณารอ 5 นาทีแล้วลองใหม่'

    # ถ้ายังไม่ถูกบล็อกและมีการ submit ฟอร์ม
    if request.method == "POST" and not blocked:
        user_id = request.POST.get('user_id')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')

        user = authenticate(request, username=user_id, password=password)

        if user is not None:
            # Log the user in
            login(request, user)

            # Set session expiration based on "Remember Me"
            if remember_me:
                request.session.set_expiry(1209600)  # 2 weeks
            else:
                request.session.set_expiry(0)  # Browser close to expire the session

            # Set session data based on user type
            if user.is_superuser:  # Check if the user is a superuser
                request.session['user_type'] = 'superuser'
                return redirect('home')  # Redirect superusers to the admin dashboard
            elif Student.objects.filter(id=user_id, delete_status='not_deleted').exists():
                request.session['user_type'] = 'student'
                request.session['student_id'] = user_id
                return redirect('home')  # Redirect students to the home page
            elif Teacher.objects.filter(id=user_id).exists():
                request.session['user_type'] = 'teacher'
                request.session['teacher_id'] = user_id
                return redirect('home')  # Redirect teachers to the home page
            else:
                # Log the user out if the category is unknown
                logout(request)
                return render(request, 'auth/login.html', {
                    'error_message': 'ไม่สามารถเข้าสู่ระบบได้',  # "Unable to log in"
                })
        else:
            # Authentication failed
            return render(request, 'auth/login.html', {
                'error_message': 'ไอดีผู้ใช้หรือรหัสผ่านไม่ถูกต้อง',  # "Incorrect user ID or password"
            })

    return render(request, 'auth/login.html', {
        'error_message': error_message,
        'blocked': blocked,
    })

# Auth
def logout_view(request):
    # Log the user out
    logout(request)
    
    # Redirect to login page after logout
    return redirect('login_view')  # Replace 'login' with the name of your login URL

#Home page
def Home(request):

    # Check if 'user_type' is in the session
    user_type = request.session.get('user_type')
    
    if not user_type:
        return redirect('login_view')  # Redirect to login if no user_type is found in session


    if user_type == 'student' and 'student_id' in request.session:
        user_id = request.session['student_id']
        
        # Get student data
        try:
            student = Student.objects.get(id=user_id)
        except Student.DoesNotExist:
            return redirect('login_view')  # Redirect if student not found

        # Get student statistics
        students = Student.objects.filter(delete_status='not_deleted')
        male_students = students.filter(gender='ชาย').count()
        female_students = students.filter(gender='หญิง').count()
        orphans = students.filter(special_status='เด็กกำพร้า').count()
        underprivileged = students.filter(special_status='เด็กยากไร้').count()
        disabled = students.filter(special_status='เด็กพิการ').count()
        new_muslims = students.filter(special_status='เด็กมุอัลลัฟ').count()

        context = {
            'user_type': user_type,
            'total_students': students.count(),
            'male_students': male_students,
            'female_students': female_students,
            'orphans': orphans,
            'underprivileged': underprivileged,
            'disabled': disabled,
            'new_muslims': new_muslims,
            'student': student
        }
        return render(request, 'student_home.html', context)
    
    elif user_type in ['teacher', 'superuser']:
        
        # Get student statistics
        students = Student.objects.filter(delete_status='not_deleted')
        male_students = students.filter(gender='ชาย').count()
        female_students = students.filter(gender='หญิง').count()
        orphans = students.filter(special_status='เด็กกำพร้า').count()
        underprivileged = students.filter(special_status='เด็กยากไร้').count()
        disabled = students.filter(special_status='เด็กพิการ').count()
        new_muslims = students.filter(special_status='เด็กมุอัลลัฟ').count()

        context = {
            'user_type': user_type,
            'total_students': students.count(),
            'male_students': male_students,
            'female_students': female_students,
            'orphans': orphans,
            'underprivileged': underprivileged,
            'disabled': disabled,
            'new_muslims': new_muslims,
        }
        return render(request, 'teacher_home.html', context)
    else:
        return redirect('login_view')  # Redirect for unauthorized access
    
#students input
def Input_Profile(request, student_id=None):
    # Check if 'user_type' is in the session
    user_type = request.session.get('user_type')
    
    # If no user_type is found, redirect to login
    if not user_type:
        return redirect('login_view')  # Redirect to login if user_type is not in session
    
    provinces = Province.objects.all()
    schools = School.objects.all()
    levels = Level.objects.all()

    student = None
    father, mother, guardian, current_study = None, None, None, None

    if student_id:
        student = get_object_or_404(Student, id=student_id)
        father = Father.objects.filter(student=student).first()
        mother = Mother.objects.filter(student=student).first()
        guardian = Guardian.objects.filter(student=student).first()
        current_study = CurrentStudy.objects.filter(student=student).first()
        

    if request.method == 'POST':
        
        def parse_date(date_str):
            try:
                day, month, year = map(int, date_str.split('/'))
                if year > 2500:
                    year -= 543  # แปลง พ.ศ. → ค.ศ.
                return datetime(year, month, day).date()
            except (ValueError, TypeError):
                return None

        def safe_int(value):
            try:
                return int(value)
            except (ValueError, TypeError):
                return None

        def handle_address(prefix):
            address_data = {
                'house_number': request.POST.get(f'{prefix}-house-number'),
                'street': request.POST.get(f'{prefix}-street'),
                'moo': safe_int(request.POST.get(f'{prefix}-village')),
                'province_id': request.POST.get(f'{prefix}-province'),
                'district_id': request.POST.get(f'{prefix}-district'),
                'subdistrict_id': request.POST.get(f'{prefix}-subdistrict'),
                'zipcode': request.POST.get(f'{prefix}-zipcode'),
            }
            return Address.objects.create(**address_data)

        # Student Data
        student_data = {
            'first_name': request.POST.get('student-first-name'),
            'last_name': request.POST.get('student-last-name'),
            'english_first_name': request.POST.get('student-english-first-name'),
            'english_last_name': request.POST.get('student-english-last-name'),
            'arabic_first_name': request.POST.get('student-arabic-first-name'),
            'arabic_last_name': request.POST.get('student-arabic-last-name'),
            'date_of_birth': parse_date(request.POST.get('student-dob')),
            'id_number': request.POST.get('student-id-number', '').replace(' ', ''),
            'special_status': request.POST.get('special-status'),
            'gender': request.POST.get('gender'),
        }

        # Validate ID number
        if not student_data['id_number'].isdigit() or len(student_data['id_number']) != 13:
            return JsonResponse({'error': 'Invalid Thai ID number. It must contain exactly 13 digits.'}, status=400)

        if student:
            for field, value in student_data.items():
                setattr(student, field, value)
            if 'profile_picture' in request.FILES:
                student.profile_picture = request.FILES['profile_picture']
            student.save()
        else:
            student = Student.objects.create(
                **student_data,
                profile_picture=request.FILES.get('profile_picture', None)
            )

        # Current Study
        school_id = request.POST.get('student-school')
        level_id = request.POST.get('student-level')
        student_school = get_object_or_404(School, id=school_id) if school_id else None
        level = Level.objects.get(id=level_id) if level_id else None

        if current_study:
            current_study.school = student_school
            current_study.level = level
            current_study.save()
        else:
            if student_school:
                CurrentStudy.objects.create(student=student, school=student_school, level=level)

        # Family Data
        for role, model in [('father', Father), ('mother', Mother), ('guardian', Guardian)]:
            address = handle_address(role)
            data = {
                'first_name': request.POST.get(f'{role}-first-name'),
                'last_name': request.POST.get(f'{role}-last-name'),
                'date_of_birth': parse_date(request.POST.get(f'{role}-dob')),
                'phone_number': request.POST.get(f'{role}-phone'),
                'occupation': request.POST.get(f'{role}-occupation'),
                'workplace': request.POST.get(f'{role}-workplace'),
                'income': safe_int(request.POST.get(f'{role}-income')),
                'address': address,
            }
            if role == 'guardian':
                data['relationship_with_student'] = request.POST.get('relationship')
            model.objects.update_or_create(student=student, defaults=data)

        return redirect('profile', student_id=student.id)

    return render(request, 'inputdata/in_profile.html', {
        'levels': levels,
        'provinces': provinces,
        'schools': schools,
        'student': student,
        'father': father,
        'mother': mother,
        'guardian': guardian,
        'current_study': current_study
    })

def Profile(request, student_id=None):
    # Check if 'user_type' is in the session
    user_type = request.session.get('user_type')
    
    # If no user_type is found, redirect to login
    if not user_type:
        return redirect('login_view')  # Redirect to login if user_type is not in session
    
    if not student_id:
        return redirect('login_view')  # Redirect to login if student_id is not provided

    # Get the student instance or return 404 if not found
    student = get_object_or_404(Student, id=student_id)
    
    # Fetch related data
    current_study = CurrentStudy.objects.filter(student=student).first()
    father = Father.objects.filter(student=student).first()
    mother = Mother.objects.filter(student=student).first()
    guardian = Guardian.objects.filter(student=student).first()

    # Pass the data to the template
    context = {
        'student': student,
        'current_study': current_study,
        'father': father,
        'mother': mother,
        'guardian': guardian,
    }

    return render(request, 'student/profile.html', context)

def Student_Rp(request):
    # Check if 'user_type' is in the session
    user_type = request.session.get('user_type')
    if not user_type:
        return redirect('login_view')
    if user_type == 'student':
        return redirect('home')

    # รับค่าฟิลเตอร์
    search = request.GET.get('search', '').strip()
    school = request.GET.get('school')
    level = request.GET.get('level')
    academic_year = request.GET.get('academic_year')
    gender = request.GET.get('gender')
    special_status = request.GET.get('special_status')
    action = request.GET.get('action')
    items_per_page = int(request.GET.get('items_per_page', 10))  # รับค่าจำนวนข้อมูลต่อหน้า (ค่าเริ่มต้นคือ 10)

    # Query นักเรียน
    students = Student.objects.filter(current_study__isnull=False, delete_status='not_deleted')

    if search:
        students = students.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search)
        )
    if school:
        students = students.filter(current_study__school__id=school)
    if level:
        students = students.filter(current_study__level__id=level)
    if academic_year:
        students = students.filter(current_study__current_semester__year=academic_year)
    if gender:
        students = students.filter(gender=gender)
    if special_status:
        students = students.filter(special_status=special_status)

    if action == 'download':
        return download_students_pdf(students)

    # Pagination
    students = students.order_by('id')  # หรือ 'first_name', 'current_study__level__name' ตามที่ต้องการ
    paginator = Paginator(students, items_per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # ดึงข้อมูลสำหรับตัวเลือก
    levels = Level.objects.all()
    schools = School.objects.all()
    academic_years = CurrentSemester.objects.values_list('year', flat=True).distinct()

    # สร้าง URL ที่รวมฟิลเตอร์ทั้งหมด
    query_params = request.GET.copy()
    query_params.pop('page', None)  # ลบพารามิเตอร์ page ออกเพื่อไม่ให้ URL ซ้ำกัน
    filter_params = query_params.urlencode()  # สร้าง URL ของพารามิเตอร์ที่เหลือ

    context = {
        'students': page_obj,
        'total_students': students.count(),
        'male_students': students.filter(gender='ชาย').count(),
        'female_students': students.filter(gender='หญิง').count(),
        'levels': levels,
        'schools': schools,
        'academic_years': academic_years,
        'current_filters': {
            'search': search,
            'school': school,
            'level': level,
            'academic_year': academic_year,
            'gender': gender,
            'special_status': special_status,
            'items_per_page': items_per_page,
        },
        'filter_params': filter_params,  # ส่งค่าพารามิเตอร์ฟิลเตอร์
    }

    return render(request, 'student/sp_student.html', context)


def download_students_pdf(students):
    # Check if there are any students
    if not students.exists():
        student_data = [['ไม่มีข้อมูล']]
    else:
        student_data = [
            ['ลำดับ', 'ชื่อ', 'นามสกุล', 'เพศ', 'โรงเรียน', 'สถานะพิเศษ'],
        ]
        for i, student in enumerate(students, start=1):
            student_data.append([
                i,
                student.first_name,
                student.last_name,
                student.gender or 'ไม่มีข้อมูล',
                student.current_study.school.name if student.current_study else 'ไม่มีข้อมูล',
                student.special_status or 'ไม่มีข้อมูล',
            ])


    
    # Dynamically determine unique values from the queryset
    unique_genders = students.values_list('gender', flat=True).distinct()
    unique_special_statuses = students.values_list('special_status', flat=True).distinct()
    unique_schools = students.values_list('current_study__school__name', flat=True).distinct()
    unique_levels = students.values_list('current_study__level__name', flat=True).distinct()
    unique_academic_years = students.values_list('current_study__current_semester__year', flat=True).distinct()

    # Function to handle the display logic based on the number of unique values
    def get_unique_text(unique_values, default_text):
        if len(unique_values) == 1:
            return unique_values[0]
        return default_text

    # Gender and Special Status Text
    gender_text = get_unique_text(unique_genders, "ทุกเพศ")
    status_text = get_unique_text(unique_special_statuses, "ทุกสถานะพิเศษ")
    # School, Level, and Academic Year Text
    school_name = get_unique_text(unique_schools, "ทุกโรงเรียน")
    level_name = get_unique_text(unique_levels, "ทุกชั้น")
    academic_year_text = get_unique_text(unique_academic_years, "ทุกปี")
    academic_year_text_thai = convert_year_to_thai(academic_year_text)
    header_info = f"ชั้น: {level_name} | ปีการศึกษา: {academic_year_text_thai} | เพศ: {gender_text} | สถานะพิเศษ: {status_text}"
    

    # Create PDF Response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="students_report_{academic_year_text_thai or "all"}.pdf"'


    # Create PDF Document (A4 Landscape)
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )

    # Logo
    logo_path = "static/images/logo.ico"
    logo = Image(logo_path, width=1 * inch, height=1 * inch)

    # School Title
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = 'THSarabunNew'
    styles['Normal'].fontSize = 20
    styles['Normal'].alignment = 1
    school_paragraph = Paragraph(f"<b>{school_name}</b>", styles['Normal'])

    # Info Paragraph
    # Custom Style for Info Paragraph
    custom_style = styles['Normal'].clone('CustomNormal')
    custom_style.fontName = 'THSarabunNew'
    custom_style.fontSize = 18
    custom_style.leading = 19  # เพิ่มระยะห่างระหว่างบรรทัด
    custom_style.alignment = 1

    # Info Paragraph
    info_paragraph = Paragraph(header_info, custom_style)

    # Header Table Layout
    header_table_data = [
        [logo, school_paragraph, info_paragraph]
    ]
    header_table = Table(
        header_table_data,
        colWidths=[2 * inch, 5 * inch, 3 * inch]
    )
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.grey),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (1, 0), (-1, -1), 15),
    ]))

    # Add spacing
    spacer = Spacer(1, 0.3 * inch)

    # Student Table
    table = Table(
        student_data,
        colWidths=[50, 150, 150, 100, 250, 100]
    )
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'THSarabunNew'),
        ('FONTSIZE', (0, 0), (-1, 0), 16),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),

        ('FONTNAME', (0, 1), (-1, -1), 'THSarabunNew'),
        ('FONTSIZE', (0, 1), (-1, -1), 14),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 12),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    # Build PDF
    elements = [header_table, spacer, table]
    doc.build(elements)

    return response

#grade input
def student_marks_view(request):
    user_type = request.session.get('user_type')
    current_semester = CurrentSemester.objects.first()

    current_thai_year = datetime.now().year + 543
    thai_years_from_db = StudentHistory.objects.values_list('academic_year', flat=True).distinct()
    clean_years = {str(int(y)) for y in thai_years_from_db if str(y).isdigit()}
    extra_years = {str(current_thai_year + offset) for offset in range(-3, 1)}
    years = sorted(clean_years.union(extra_years), reverse=True)

    academic_year = request.GET.get('academic_year')
    if not academic_year:
        academic_year = str(current_semester.year + 543) if current_semester else str(current_thai_year)

    try:
        academic_year_int = int(academic_year) - 543
    except:
        academic_year_int = datetime.now().year

    if not user_type:
        return redirect('login_view')
    if user_type == 'student':
        return redirect('home')

    school_name = request.GET.get('school')
    level_name = request.GET.get('level')

    schools = School.objects.all()
    levels = Level.objects.all()
    students = []
    subjects = []
    student_marks_data = []

    if school_name and level_name:
        students_query = CurrentStudy.objects.filter(
            current_semester=current_semester,
            school__name__iexact=school_name,
            level__name__iexact=level_name,
        ).select_related('student', 'level', 'school')

        students = list(students_query)

        subjects = SubjectToStudy.objects.filter(
            level__name__iexact=level_name
        ).select_related('subject')

        for student in students:
            marks_row = {'student': student.student}
            for subject in subjects:
                mark_obj = StudentMarkForSubject.objects.filter(
                    student=student.student,
                    subject_to_study=subject,
                    academic_year=academic_year_int
                ).first()
                marks_row[subject.subject.id] = mark_obj.marks_obtained if mark_obj else ''
            student_marks_data.append(marks_row)

    if request.method == 'POST':
        academic_year = request.POST.get('academic_year') or academic_year
        try:
            academic_year_int = int(academic_year) - 543
        except:
            academic_year_int = datetime.now().year

        students_query = CurrentStudy.objects.filter(
            current_semester=current_semester,
            school__name__iexact=school_name,
            level__name__iexact=level_name,
            student__delete_status='not_deleted'
        ).select_related('student', 'level', 'school')

        subjects = SubjectToStudy.objects.filter(
            level__name__iexact=level_name
        ).select_related('subject')

        for student in students_query:
            student_subject_marks = {}
            total_marks = 0
            obtained_marks = 0

            for subject in subjects:
                field_name = f"marks_{student.student.id}_{subject.subject.id}"
                marks = request.POST.get(field_name)

                if marks:
                    try:
                        marks = int(marks)
                        student_subject_marks[subject.subject.name] = marks
                        total_marks += subject.subject.total_marks
                        obtained_marks += marks

                        StudentMarkForSubject.objects.update_or_create(
                            student=student.student,
                            subject_to_study=subject,
                            category=subject.subject.category,
                            academic_year=academic_year_int,
                            defaults={'marks_obtained': marks},
                        )
                    except ValueError:
                        return render(request, 'inputdata/ingr_student.html', {
                            'error': f"Invalid marks for {student.student.first_name} in {subject.subject.name}.",
                            'schools': schools,
                            'levels': levels,
                            'students': students,
                            'subjects': subjects,
                        })

            grade_percentage = (obtained_marks / total_marks) * 100 if total_marks > 0 else 0

            StudentHistory.objects.update_or_create(
                student_id=student.student.id,
                student_name=f"{student.student.first_name} {student.student.last_name}",
                school_name=student.school.name,
                level_name=student.level.name,
                academic_year=academic_year_int,
                defaults={
                    'total_marks': total_marks,
                    'obtained_marks': obtained_marks,
                    'grade_percentage': grade_percentage,
                    'subject_marks': student_subject_marks,
                    'pass_or_fail': "ผ่าน" if grade_percentage >= 50 else "ไม่ผ่าน"
                }
            )

        query_params = {
            'school': school_name,
            'level': level_name,
            'academic_year': academic_year,
        }
        query_params = {k: v for k, v in query_params.items() if v}
        redirect_url = f"{reverse('gr_student')}?{urlencode(query_params)}"
        return HttpResponseRedirect(redirect_url)

    context = {
        'user_type': user_type,
        'schools': schools,
        'levels': levels,
        'students': students,
        'subjects': subjects,
        'current_semester': current_semester,
        'student_marks_data': student_marks_data,
        'academic_year': academic_year,
        'years': years,
    }

    return render(request, 'inputdata/ingr_student.html', context)

#grade output
def GR_Student(request):
    user_type = request.session.get('user_type')
    if not user_type:
        return redirect('login_view')
    if user_type == 'student':
        return redirect('home')

    schools = School.objects.all()
    levels = Level.objects.all()
    current_semester = CurrentSemester.objects.first()
    current_year = int(current_semester.year) if current_semester else datetime.now().year + 543

    academic_years = StudentHistory.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')

    school_name = request.GET.get('school')
    level_name = request.GET.get('level')
    academic_year = request.GET.get('academic_year') or str(current_year)

    histories = StudentHistory.objects.all()
    if school_name:
        histories = histories.filter(school_name=school_name)
    if level_name:
        histories = histories.filter(level_name=level_name)
    if academic_year:
        histories = histories.filter(academic_year=academic_year)

    # เก็บชื่อวิชาที่นักเรียนมีใน subject_marks
    used_subject_names = set()
    for history in histories:
        if history.subject_marks:
            used_subject_names.update(history.subject_marks.keys())

    # ดึงวิชา และแยกตามประเภท (category)
    subject_study_qs = SubjectToStudy.objects.filter(level__name=level_name)
    practical_subjects = [s.subject.name for s in subject_study_qs if s.subject.category == 2]
    theory_subjects = [s.subject.name for s in subject_study_qs if s.subject.category == 1]
    subject_totals = {s.subject.name: s.subject.total_marks for s in subject_study_qs}

    context = {
        'user_type': user_type,
        'schools': schools,
        'levels': levels,
        'academic_years': academic_years,
        'students': histories,
        'academic_year': academic_year,
        'school_name': school_name,
        'level_name': level_name,
        'practical_subjects': practical_subjects,
        'theory_subjects': theory_subjects,
        'subject_totals': subject_totals,
    }
    return render(request, 'student/gr_student.html', context)

def download_student_results_excel(request):
    school_name = request.GET.get('school')
    level_name = request.GET.get('level')
    academic_year = request.GET.get('academic_year')

    histories = StudentHistory.objects.all()
    if school_name:
        histories = histories.filter(school_name=school_name)
    if level_name:
        histories = histories.filter(level_name=level_name)
    if academic_year:
        histories = histories.filter(academic_year=academic_year)

    theory_subjects = list(Subject.objects.filter(category=1).values_list('name', flat=True))
    practical_subjects = list(Subject.objects.filter(category=2).values_list('name', flat=True))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ผลการเรียน"

    # --- โลโก้ตรงกลาง (แถว 1–2) ---
    logo_path = 'static/images/logo.png'
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 80
        logo.height = 80
        ws.add_image(logo, 'F1')  # ประมาณกลางหน้า

    academic_year_thai = str(int(academic_year) + 543) if academic_year and academic_year.isdigit() else "ทุกปี"
    level_display = level_name or "ทุกระดับชั้น"
    school_display = school_name or "ทุกโรงเรียน"
    province = "จังหวัดกระบี่"

    # --- คำนวณจำนวนคอลัมน์ทั้งหมด ---
    last_col = 2 + len(theory_subjects) + len(practical_subjects) + 3
    last_col_letter = get_column_letter(last_col)

    # --- แถว 5: ชื่อสมาคมฯ ---
    ws.merge_cells(f'A5:{last_col_letter}5')
    ws['A5'] = f'สมาคมวิชาการศาสนาอิสลามภาคฟัรฎูกิฟายะห์ ปีการศึกษา {academic_year_thai} ระดับชั้นปี {level_display}'
    ws['A5'].font = Font(size=14, bold=True)
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')

    # --- แถว 6: โรงเรียน จังหวัด ---
    ws.merge_cells(f'A6:{last_col_letter}6')
    ws['A6'] = f'{school_display} {province}'
    ws['A6'].font = Font(size=12)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')

    start_row = 8

    # --- หัวตารางหลัก ---
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
    ws.cell(row=start_row, column=1).value = 'ลำดับ'
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row+1, end_column=2)
    ws.cell(row=start_row, column=2).value = 'ชื่อ-สกุล'

    col_index = 3
    if theory_subjects:
        end_col = col_index + len(theory_subjects) - 1
        ws.merge_cells(start_row=start_row, start_column=col_index, end_row=start_row, end_column=end_col)
        ws.cell(row=start_row, column=col_index).value = 'ภาคทฤษฎี'
        col_index = end_col + 1

    if practical_subjects:
        end_col = col_index + len(practical_subjects) - 1
        ws.merge_cells(start_row=start_row, start_column=col_index, end_row=start_row, end_column=end_col)
        ws.cell(row=start_row, column=col_index).value = 'ภาคปฏิบัติ'
        col_index = end_col + 1

    ws.merge_cells(start_row=start_row, start_column=col_index, end_row=start_row, end_column=col_index + 2)
    ws.cell(row=start_row, column=col_index).value = 'สรุปผล'

    # --- Subheader: วิชา + สรุป ---
    col = 3
    for subject in theory_subjects:
        ws.cell(row=start_row+1, column=col).value = subject
        ws.cell(row=start_row+1, column=col).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
        col += 1
    for subject in practical_subjects:
        ws.cell(row=start_row+1, column=col).value = subject
        ws.cell(row=start_row+1, column=col).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
        col += 1
    for i, title in enumerate(['คะแนนรวม', 'คิดเป็นร้อยละ', 'ผลตัดสิน']):
        ws.cell(row=start_row+1, column=col+i).value = title
        ws.cell(row=start_row+1, column=col+i).alignment = Alignment(horizontal='center', vertical='center')

    # --- หัวตาราง font + alignment ---
    for cell in ws[start_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[start_row+1]:
        cell.font = Font(bold=True)

    # --- ข้อมูลนักเรียน ---
    for idx, student in enumerate(histories, start=1):
        row = start_row + 1 + idx
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = student.student_name
        col = 3
        for subject in theory_subjects:
            ws.cell(row=row, column=col).value = student.subject_marks.get(subject, '-') if student.subject_marks else '-'
            col += 1
        for subject in practical_subjects:
            ws.cell(row=row, column=col).value = student.subject_marks.get(subject, '-') if student.subject_marks else '-'
            col += 1
        ws.cell(row=row, column=col).value = student.obtained_marks or 0
        ws.cell(row=row, column=col+1).value = round(student.grade_percentage, 1) if student.grade_percentage is not None else '-'
        ws.cell(row=row, column=col+2).value = student.pass_or_fail or '-'

        for c in range(1, col + 3):
            ws.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')

    # --- ใส่ border ให้ทุก cell ที่ใช้ ---
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = border


    for idx in range(1, last_col + 1):
        col_letter = get_column_letter(idx)
        if idx == 1:
            ws.column_dimensions[col_letter].width = 7   # ลำดับ
        elif idx == 2:
            ws.column_dimensions[col_letter].width = 22  # ชื่อ
        elif 3 <= idx < 3 + len(theory_subjects) + len(practical_subjects):
            ws.column_dimensions[col_letter].width = 5   # วิชา
        else:
            ws.column_dimensions[col_letter].width = 13  # คะแนนรวม ฯลฯ


    # --- ชื่อไฟล์ ---
    filename = f"ผลการเรียน_{school_display}_{academic_year_thai}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    encoded_filename = quote(filename)
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"    
    wb.save(response)
    return response

#grade output
def student_Results(request, student_id):
    user_type = request.session.get('user_type')
    if not user_type:
        return redirect('login_view')

    student = get_object_or_404(Student, id=student_id)
    current_study = CurrentStudy.objects.filter(student=student).first()
    current_semester = CurrentSemester.objects.first()

    academic_years = StudentHistory.objects.filter(student_id=student.id) \
        .values_list('academic_year', flat=True).distinct().order_by('-academic_year')

    selected_academic_year = request.GET.get('academic_year')
    selected_category = request.GET.get('category')  # 1 = ทฤษฎี, 2 = ปฏิบัติ

    if not selected_academic_year:
        if academic_years:
            selected_academic_year = academic_years.first()
        elif current_semester:
            selected_academic_year = current_semester.year

    # ดึงข้อมูลตามปีและประเภทวิชา
    cat1_data = StudentHistory.objects.filter(
        student_id=student.id,
        academic_year=selected_academic_year,
        category=1
    ).first()

    cat2_data = StudentHistory.objects.filter(
        student_id=student.id,
        academic_year=selected_academic_year,
        category=2
    ).first()

    # ดึงข้อมูลรายวิชาตามประเภท
    if selected_category == '1':
        subjects_cat1 = cat1_data.get_subject_data(category=1) if cat1_data else []
        subjects_cat2 = []
    elif selected_category == '2':
        subjects_cat1 = []
        subjects_cat2 = cat2_data.get_subject_data(category=2) if cat2_data else []
    else:
        subjects_cat1 = cat1_data.get_subject_data(category=1) if cat1_data else []
        subjects_cat2 = cat2_data.get_subject_data(category=2) if cat2_data else []

    # ถ้ากดปุ่มดาวน์โหลด PDF
    if 'download_pdf' in request.GET:
        def format_data(data):
            return [(s['name'], s['marks'], f"{round(s['percentage'])}%", s['status']) for s in data]

        return download_result_pdfs(
            student=student,
            semester_1_data=format_data(subjects_cat1),
            semester_2_data=format_data(subjects_cat2),
            academic_year=convert_to_thai_year(str(selected_academic_year)),
            student_name=f"{student.first_name}_{student.last_name}",
            school_name=current_study.school.name if current_study and current_study.school else "-",
            level_name=current_study.level.name if current_study and current_study.level else "-",
            selected_semester=None  # ไม่มี semester แล้ว
        )

    context = {
        'student': student,
        'current_study': current_study,
        'current_semester': current_semester,
        'selected_academic_year': selected_academic_year,
        'selected_category': selected_category,
        'academic_years': academic_years,
        'subjects_cat1': subjects_cat1,
        'subjects_cat2': subjects_cat2,
    }

    return render(request, 'student/student_results.html', context)


def download_result_pdfs(
    student,
    semester_1_data,
    semester_2_data,
    academic_year,
    student_name,
    school_name,
    level_name,
    selected_semester=None,
):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student_name}_results_{academic_year}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )

    # Styles
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = 'THSarabunNew'
    styles['Normal'].fontSize = 18

    title_style = ParagraphStyle(
        name='Title',
        fontName='THSarabunNew',
        fontSize=16,
        alignment=1
    )
    sub_title_style = ParagraphStyle(
        name='SubTitle',
        fontName='THSarabunNew',
        fontSize=16,
        alignment=1
    )
    normal_style = ParagraphStyle(
        name='Normal',
        fontName='THSarabunNew',
        fontSize=16,
        alignment=0
    )

    # Logo
    logo_path = "static/images/logo.ico"
    logo = Image(logo_path, width=80, height=80)
    logo.hAlign = 'CENTER'

    # Header
    header = Paragraph("สมาคมคุรุสัมพันธ์อิสลามแห่งประเทศไทย ประจำหน่วยสอบที่ 80", title_style)
    school_name_paragraph = Paragraph(f"โรงเรียน {school_name}", sub_title_style)
    prefix = 'เด็กชาย' if student.gender == 'ชาย' else 'เด็กหญิง' if student.gender == 'หญิง' else ''
    student_info = Paragraph(f"ผลการเรียนของ {prefix} {student_name} {level_name}", normal_style)
    academic_year_text = Paragraph(f"ประจำปีการศึกษา {academic_year}", normal_style)

    # Tables
    def generate_table(data):
        if not data:
            return Paragraph("ยังไม่มีข้อมูลสำหรับภาคนี้", sub_title_style)
        table_data = [['วิชา', 'คะแนน', 'เกรด', 'ผ่าน/ไม่ผ่าน']] + data
        table = Table(table_data, colWidths=[200, 100, 100, 100])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'THSarabunNew'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, 0), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        return table

    # Elements
    elements = [
        logo,
        Spacer(1, 10),
        header,
        Spacer(1, 20),
        school_name_paragraph,
        Spacer(1, 20),
        student_info,
        Spacer(1, 10),
        academic_year_text,
        Spacer(1, 20),
    ]

    if selected_semester == '1':
        if semester_1_data:
            elements += [
                Paragraph("ผลการเรียนภาคทฤษฎี", title_style),
                Spacer(1, 20),
                generate_table(semester_1_data)
            ]
    elif selected_semester == '2':
        if semester_2_data:
            elements += [
                Paragraph("ผลการเรียนภาคปฏิบัติ", title_style),
                Spacer(1, 20),
                generate_table(semester_2_data)
            ]
    else:
        if semester_1_data:
            elements += [
                Paragraph("ผลการเรียนภาคทฤษฎี", title_style),
                Spacer(1, 20),
                generate_table(semester_1_data),
                Spacer(1, 15),
            ]
        if semester_2_data:
            elements += [
                Paragraph("ผลการเรียนภาคปฏิบัติ", title_style),
                Spacer(1, 20),
                generate_table(semester_2_data),
            ]

    # สร้าง PDF
    doc.build(elements)
    return response



# error
# 403 - Forbidden
def test_403_view(request):
    raise PermissionDenied("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")

# 404 - Not Found
def test_404_view(request):
    raise Http404("ไม่พบหน้านี้ในระบบ")

# 500 - Server Error
def test_500_view(request):
    1 / 0  # ทำให้เกิด ZeroDivisionError เพื่อ simulate server error

def custom_csrf_failure(request, reason=""):
    return render(request, 'csrf_403.html', status=403)


